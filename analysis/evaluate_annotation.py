"""Recalculate inter-annotator agreement and held-out model performance."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--agreement",
        type=Path,
        default=Path("data/inter_annotator_agreement.xlsx"),
    )
    parser.add_argument("--agreement-sheet", default="2022 Annotation Comparison")
    parser.add_argument(
        "--test-set",
        type=Path,
        default=Path(
            "semantic_annotation/training_data/test_set_84.xlsx"
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("analysis/results/annotation_evaluation.json"),
    )
    parser.add_argument("--force", action="store_true")
    return parser.parse_args()


def text(value: object) -> str:
    return re.sub(r"\s+", " ", "" if value is None else str(value)).strip()


def cohen_kappa(left: list[str], right: list[str]) -> dict[str, float | int]:
    if len(left) != len(right) or not left:
        raise ValueError("Kappa inputs must have equal, nonzero length")
    n = len(left)
    observed_count = sum(a == b for a, b in zip(left, right))
    left_counts = Counter(left)
    right_counts = Counter(right)
    labels = sorted(set(left) | set(right))
    expected = sum(
        left_counts[label] * right_counts[label] for label in labels
    ) / (n * n)
    observed = observed_count / n
    kappa = (observed - expected) / (1 - expected)
    return {
        "n": n,
        "observedAgreement": observed,
        "expectedAgreement": expected,
        "cohenKappa": kappa,
    }


def classification_metrics(
    truth: list[str], prediction: list[str]
) -> dict[str, object]:
    if len(truth) != len(prediction) or not truth:
        raise ValueError("Classification inputs must have equal, nonzero length")
    labels = sorted(set(truth) | set(prediction))
    matrix = {
        actual: {
            predicted: sum(
                a == actual and p == predicted
                for a, p in zip(truth, prediction)
            )
            for predicted in labels
        }
        for actual in labels
    }
    rows = []
    for label in labels:
        tp = matrix[label][label]
        support = sum(matrix[label].values())
        predicted_count = sum(matrix[actual][label] for actual in labels)
        precision = tp / predicted_count if predicted_count else 0.0
        recall = tp / support if support else 0.0
        f1 = (
            2 * precision * recall / (precision + recall)
            if precision + recall
            else 0.0
        )
        rows.append(
            {
                "category": label,
                "precision": precision,
                "recall": recall,
                "f1": f1,
                "support": support,
            }
        )
    n = len(truth)
    accuracy = sum(a == p for a, p in zip(truth, prediction)) / n
    macro = {
        key: sum(row[key] for row in rows) / len(rows)
        for key in ("precision", "recall", "f1")
    }
    weighted = {
        key: sum(row[key] * row["support"] for row in rows) / n
        for key in ("precision", "recall", "f1")
    }
    return {
        "n": n,
        "accuracy": accuracy,
        "labels": labels,
        "categoryMetrics": rows,
        "macroAverage": macro,
        "weightedAverage": weighted,
        "confusionMatrix": matrix,
    }


def load_column_pair(
    path: Path,
    sheet_name: str,
    left_column: str,
    right_column: str,
) -> tuple[list[str], list[str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[sheet_name]
    left_index = column_index_from_string(left_column) - 1
    right_index = column_index_from_string(right_column) - 1
    left: list[str] = []
    right: list[str] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        left_value = text(row[left_index])
        right_value = text(row[right_index])
        if not left_value and not right_value:
            continue
        if not left_value or not right_value:
            raise ValueError(
                f"Incomplete label pair in {path.name}, sheet {sheet_name}"
            )
        left.append(left_value)
        right.append(right_value)
    return left, right


def main() -> int:
    args = parse_args()
    output_path = args.output.resolve()
    if output_path.exists() and not args.force:
        raise FileExistsError(
            f"Output already exists: {output_path}. Pass --force to replace it."
        )

    agreement_path = args.agreement.resolve()
    function_left, function_right = load_column_pair(
        agreement_path, args.agreement_sheet, "K", "L"
    )
    depth_left, depth_right = load_column_pair(
        agreement_path, args.agreement_sheet, "M", "N"
    )

    test_path = args.test_set.resolve()
    test_workbook = load_workbook(test_path, data_only=True, read_only=True)
    test_sheet = test_workbook.active
    rows = list(test_sheet.iter_rows(min_row=2, values_only=True))
    function_truth = [text(row[4]) for row in rows]
    function_prediction = [text(row[5]) for row in rows]
    depth_truth = [text(row[6]) for row in rows]
    depth_prediction = [text(row[7]) for row in rows]

    output = {
        "inputs": {
            "agreementFile": args.agreement.as_posix(),
            "agreementSheet": args.agreement_sheet,
            "agreementColumns": {
                "citationFunction": ["K", "L"],
                "citationDepth": ["M", "N"],
            },
            "testSetFile": args.test_set.as_posix(),
            "testSetColumns": {
                "citationFunction": ["E", "F"],
                "citationDepth": ["G", "H"],
            },
        },
        "interAnnotatorAgreement": {
            "citationFunction": cohen_kappa(function_left, function_right),
            "citationDepth": cohen_kappa(depth_left, depth_right),
        },
        "heldOutModelEvaluation": {
            "citationFunction": classification_metrics(
                function_truth, function_prediction
            ),
            "citationDepth": classification_metrics(
                depth_truth, depth_prediction
            ),
        },
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(output, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "output": str(output_path),
                "agreementN": len(function_left),
                "functionKappa": output["interAnnotatorAgreement"][
                    "citationFunction"
                ]["cohenKappa"],
                "depthKappa": output["interAnnotatorAgreement"][
                    "citationDepth"
                ]["cohenKappa"],
                "testN": len(rows),
                "functionAccuracy": output["heldOutModelEvaluation"][
                    "citationFunction"
                ]["accuracy"],
                "depthAccuracy": output["heldOutModelEvaluation"][
                    "citationDepth"
                ]["accuracy"],
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
