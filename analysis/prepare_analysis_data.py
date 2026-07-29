"""Prepare the released annotation workbook for statistical analysis."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_COLUMNS = (
    "Lable",
    "Journal",
    "Year",
    "Self-citing article",
    "Self-cited article",
    "Citation content",
    "Citation position",
    "Citation distance",
    "Citation strength",
    "Citation function",
    "Citation depth",
    "Semantic similarity",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Convert Table S2 of the released annotation workbook into the "
            "record-oriented JSON consumed by run_full_analysis.py."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("data/full_annotation_results.xlsx"),
    )
    parser.add_argument("--sheet", default="Table S2")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("analysis/results/analysis_records.json"),
    )
    parser.add_argument("--expect-mentions", type=int, default=2435)
    parser.add_argument("--expect-strength-occurrences", type=int, default=1883)
    parser.add_argument("--force", action="store_true")
    return parser.parse_args()


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def main() -> int:
    args = parse_args()
    input_path = args.input.resolve()
    output_path = args.output.resolve()
    if not input_path.is_file():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")
    if output_path.exists() and not args.force:
        raise FileExistsError(
            f"Output already exists: {output_path}. Pass --force to replace it."
        )

    workbook = load_workbook(input_path, data_only=True, read_only=True)
    if args.sheet not in workbook.sheetnames:
        raise ValueError(
            f"Worksheet {args.sheet!r} not found; available: {workbook.sheetnames}"
        )
    worksheet = workbook[args.sheet]
    rows = worksheet.iter_rows(values_only=True)
    raw_headers = next(rows)
    headers = [text(value) for value in raw_headers]
    column = {header: index for index, header in enumerate(headers)}
    missing = [name for name in REQUIRED_COLUMNS if name not in column]
    if missing:
        raise ValueError(
            f"Missing required columns: {missing}. Available columns: {headers}"
        )

    records: list[dict[str, object]] = []
    article_number = 0
    strength_number = 0
    current_article_title = ""

    for source_row, row in enumerate(rows, start=2):
        raw_article_title = text(row[column["Self-citing article"]])
        if raw_article_title:
            article_number += 1
            current_article_title = raw_article_title

        raw_strength = row[column["Citation strength"]]
        has_strength = raw_strength is not None and text(raw_strength) != ""
        if has_strength:
            strength_number += 1

        label = text(row[column["Lable"]])
        raw_position = text(row[column["Citation position"]])
        position = "Discussion" if raw_position.casefold() == "discussion" else raw_position
        similarity = row[column["Semantic similarity"]]
        if similarity is None or text(similarity) == "":
            raise ValueError(f"Row {source_row}: Semantic similarity is blank")

        records.append(
            {
                "sourceRow": source_row,
                "label": label,
                "group": "G1" if label.startswith("G1") else "G2",
                "field": label.split("-", 1)[1] if "-" in label else "",
                "journal": text(row[column["Journal"]]),
                "year": row[column["Year"]],
                "articleId": f"A{article_number:04d}",
                "articleTitle": current_article_title,
                "articleTitleAnchor": raw_article_title,
                "selfCitedArticleRaw": text(row[column["Self-cited article"]]),
                "citationContent": text(row[column["Citation content"]]),
                "position": position,
                "positionRaw": raw_position,
                "distance": text(row[column["Citation distance"]]),
                "hasStrength": has_strength,
                "strengthId": f"S{strength_number:04d}" if has_strength else "",
                "strength": float(raw_strength) if has_strength else None,
                "citationFunction": text(row[column["Citation function"]]),
                "citationDepth": text(row[column["Citation depth"]]),
                "semanticSimilarity": float(similarity),
            }
        )

    if len(records) != args.expect_mentions:
        raise ValueError(
            f"Expected {args.expect_mentions} mention records, found {len(records)}"
        )
    if strength_number != args.expect_strength_occurrences:
        raise ValueError(
            "Expected "
            f"{args.expect_strength_occurrences} Citation Strength occurrences, "
            f"found {strength_number}"
        )

    payload = {
        "source": args.input.as_posix(),
        "sheet": args.sheet,
        "range": f"A1:M{len(records) + 1}",
        "headers": headers,
        "records": records,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "output": str(output_path),
                "mentions": len(records),
                "articles": article_number,
                "strengthOccurrences": strength_number,
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
