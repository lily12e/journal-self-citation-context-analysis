"""Clean citation-context text and highlight records that need manual review."""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


DEFAULT_CITATION_COLUMNS = (
    "Citation content",
    "Citation Content",
    "Citation Content(卤3)",
    "Citation Content (卤3)",
)
DEFAULT_SELF_CITED_COLUMNS = (
    "Self-cited article",
    "Self-cited Article",
    "Self-cited Article title",
    "Self-cited Article Title",
)


def clean_text(content: object) -> str:
    """Normalize whitespace and repeated dash characters."""
    if content is None or pd.isna(content):
        return ""
    text = re.sub(r"\s+", " ", str(content))
    text = re.sub(r"[-\u2013\u2014]{2,}", " ", text)
    return text.strip()


def has_multiple_years(self_cited_article: object) -> bool:
    """Return True when a self-cited reference contains multiple distinct years."""
    if not isinstance(self_cited_article, str):
        return False
    years = re.findall(r"\b(?:19|20)\d{2}[a-z]?\b", self_cited_article)
    return len(set(years)) > 1


def is_short_citation(content: object, min_length: int = 350) -> bool:
    """Return True when the cleaned citation context is shorter than min_length."""
    return len(clean_text(content)) < min_length


def has_excessive_dashes(content: object, threshold: int = 10) -> bool:
    """Return True when a citation context contains more than threshold dashes."""
    if not isinstance(content, str):
        return False
    dash_count = content.count("-") + content.count("\u2013") + content.count("\u2014")
    return dash_count > threshold


def normalize_header(value: object) -> str:
    """Normalize a header for case-insensitive matching and stray-space handling."""
    return re.sub(r"\s+", " ", str(value)).strip().casefold()


def detect_column_name(
    df: pd.DataFrame,
    candidates: Sequence[str],
    label: str,
) -> str:
    """Return the first matching column, preserving its exact workbook spelling."""
    for candidate in candidates:
        if candidate in df.columns:
            return candidate

    normalized_columns = {
        normalize_header(column): column
        for column in df.columns
    }
    for candidate in candidates:
        normalized = normalize_header(candidate)
        if normalized in normalized_columns:
            return normalized_columns[normalized]

    raise KeyError(
        f"Could not find the {label} column. Expected one of {list(candidates)}; "
        f"found {list(df.columns)}."
    )


def resolve_sheet_name(
    input_xlsx: Path,
    requested_sheet: str,
    citation_candidates: Sequence[str],
    self_cited_candidates: Sequence[str],
) -> str:
    """Resolve a sheet name or automatically select the first compatible sheet."""
    with pd.ExcelFile(input_xlsx, engine="openpyxl") as workbook:
        sheet_names = workbook.sheet_names

        if requested_sheet.casefold() == "auto":
            failures = []
            for sheet_name in sheet_names:
                headers = pd.read_excel(
                    workbook,
                    sheet_name=sheet_name,
                    nrows=0,
                )
                try:
                    detect_column_name(headers, citation_candidates, "citation content")
                    detect_column_name(headers, self_cited_candidates, "self-cited article")
                    return sheet_name
                except KeyError as exc:
                    failures.append(f"{sheet_name}: {exc}")
            raise KeyError(
                "No compatible worksheet was found. Checked: "
                + " | ".join(failures)
            )

        if requested_sheet.isdigit():
            sheet_index = int(requested_sheet)
            if sheet_index >= len(sheet_names):
                raise IndexError(
                    f"Worksheet index {sheet_index} is out of range. "
                    f"Available sheets: {sheet_names}."
                )
            return sheet_names[sheet_index]

        if requested_sheet not in sheet_names:
            raise KeyError(
                f"Worksheet '{requested_sheet}' was not found. "
                f"Available sheets: {sheet_names}."
            )
        return requested_sheet


def validate_workbook(
    input_xlsx: Path,
    requested_sheet: str = "auto",
    citation_candidates: Sequence[str] = DEFAULT_CITATION_COLUMNS,
    self_cited_candidates: Sequence[str] = DEFAULT_SELF_CITED_COLUMNS,
) -> tuple[pd.DataFrame, str, str, str]:
    """Validate the input workbook and return its data and resolved field names."""
    if not input_xlsx.is_file():
        raise FileNotFoundError(f"Input workbook does not exist: {input_xlsx}")
    if input_xlsx.suffix.casefold() != ".xlsx":
        raise ValueError(f"Input must be an .xlsx file: {input_xlsx}")

    sheet_name = resolve_sheet_name(
        input_xlsx,
        requested_sheet,
        citation_candidates,
        self_cited_candidates,
    )
    df = pd.read_excel(
        input_xlsx,
        sheet_name=sheet_name,
        engine="openpyxl",
    )
    citation_column = detect_column_name(
        df,
        citation_candidates,
        "citation content",
    )
    self_cited_column = detect_column_name(
        df,
        self_cited_candidates,
        "self-cited article",
    )
    return df, sheet_name, citation_column, self_cited_column


def clean_and_highlight(
    df: pd.DataFrame,
    citation_column: str,
    self_cited_column: str,
    worksheet,
    min_length: int,
) -> dict[str, int]:
    """Write cleaned text and apply review highlights to the selected worksheet."""
    red_fill = PatternFill(
        start_color="FF9999",
        end_color="FF9999",
        fill_type="solid",
    )
    blue_fill = PatternFill(
        start_color="99CCFF",
        end_color="99CCFF",
        fill_type="solid",
    )

    cleaned_column = f"{citation_column}_cleaned"
    cleaned_values = df[citation_column].map(clean_text)
    cleaned_column_index = len(df.columns) + 1

    existing_headers = {
        worksheet.cell(row=1, column=index).value: index
        for index in range(1, worksheet.max_column + 1)
    }
    if cleaned_column in existing_headers:
        cleaned_column_index = existing_headers[cleaned_column]
    else:
        worksheet.cell(
            row=1,
            column=cleaned_column_index,
            value=cleaned_column,
        )

    citation_column_index = df.columns.get_loc(citation_column) + 1
    self_cited_column_index = df.columns.get_loc(self_cited_column) + 1
    counts = {
        "rows": len(df),
        "multiple_years": 0,
        "short_citations": 0,
        "excessive_dashes": 0,
    }

    for row_offset, (_, row) in enumerate(df.iterrows(), start=2):
        citation_content = row.get(citation_column)
        self_cited_article = row.get(self_cited_column)

        if has_multiple_years(self_cited_article):
            worksheet.cell(
                row=row_offset,
                column=self_cited_column_index,
            ).fill = blue_fill
            counts["multiple_years"] += 1

        if is_short_citation(citation_content, min_length=min_length):
            worksheet.cell(
                row=row_offset,
                column=citation_column_index,
            ).fill = red_fill
            counts["short_citations"] += 1
        elif has_excessive_dashes(citation_content):
            worksheet.cell(
                row=row_offset,
                column=citation_column_index,
            ).fill = blue_fill
            counts["excessive_dashes"] += 1

        worksheet.cell(
            row=row_offset,
            column=cleaned_column_index,
            value=cleaned_values.iloc[row_offset - 2],
        )

    return counts


def process_workbook(
    input_xlsx: Path,
    output_xlsx: Path,
    requested_sheet: str = "auto",
    citation_candidates: Sequence[str] = DEFAULT_CITATION_COLUMNS,
    self_cited_candidates: Sequence[str] = DEFAULT_SELF_CITED_COLUMNS,
    min_length: int = 350,
    overwrite: bool = False,
) -> dict[str, int]:
    """Validate, clean, highlight, and save a copy of an annotation workbook."""
    df, sheet_name, citation_column, self_cited_column = validate_workbook(
        input_xlsx,
        requested_sheet=requested_sheet,
        citation_candidates=citation_candidates,
        self_cited_candidates=self_cited_candidates,
    )

    if input_xlsx.resolve() == output_xlsx.resolve():
        raise ValueError("The output path must differ from the input path.")
    if output_xlsx.exists() and not overwrite:
        raise FileExistsError(
            f"Output already exists: {output_xlsx}. "
            "Use --overwrite to replace it."
        )

    workbook = load_workbook(input_xlsx)
    worksheet = workbook[sheet_name]
    counts = clean_and_highlight(
        df,
        citation_column,
        self_cited_column,
        worksheet,
        min_length=min_length,
    )

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)
    return counts


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Clean citation-context text and highlight records that need "
            "manual review in an annotation workbook."
        )
    )
    parser.add_argument("input_xlsx", type=Path, help="Input .xlsx workbook.")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help=(
            "Output .xlsx workbook. Defaults to "
            "<input_stem>_checked.xlsx beside the input file."
        ),
    )
    parser.add_argument(
        "--sheet",
        default="auto",
        help=(
            "Worksheet name or zero-based index. The default, 'auto', selects "
            "the first sheet containing the required columns."
        ),
    )
    parser.add_argument(
        "--citation-column",
        help="Exact citation-content column name, if automatic detection is unsuitable.",
    )
    parser.add_argument(
        "--self-cited-column",
        help="Exact self-cited-article column name, if automatic detection is unsuitable.",
    )
    parser.add_argument(
        "--min-length",
        type=int,
        default=350,
        help="Minimum cleaned citation-context length (default: 350 characters).",
    )
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Validate the workbook, sheet, and columns without creating output.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Allow replacement of an existing output file (never the input file).",
    )
    return parser


def cli() -> None:
    args = build_parser().parse_args()
    if args.min_length < 0:
        raise ValueError("--min-length must be zero or greater.")

    citation_candidates = (
        (args.citation_column,)
        if args.citation_column
        else DEFAULT_CITATION_COLUMNS
    )
    self_cited_candidates = (
        (args.self_cited_column,)
        if args.self_cited_column
        else DEFAULT_SELF_CITED_COLUMNS
    )

    input_xlsx = args.input_xlsx
    output_xlsx = args.output or input_xlsx.with_name(
        f"{input_xlsx.stem}_checked.xlsx"
    )

    if args.validate_only:
        df, sheet_name, citation_column, self_cited_column = validate_workbook(
            input_xlsx,
            requested_sheet=args.sheet,
            citation_candidates=citation_candidates,
            self_cited_candidates=self_cited_candidates,
        )
        print(f"Validation passed: {input_xlsx}")
        print(f"Worksheet: {sheet_name}")
        print(f"Rows: {len(df)}")
        print(f"Citation column: {citation_column}")
        print(f"Self-cited article column: {self_cited_column}")
        return

    counts = process_workbook(
        input_xlsx,
        output_xlsx,
        requested_sheet=args.sheet,
        citation_candidates=citation_candidates,
        self_cited_candidates=self_cited_candidates,
        min_length=args.min_length,
        overwrite=args.overwrite,
    )
    print(f"Completed: {output_xlsx}")
    print(
        "Review flags: "
        f"{counts['short_citations']} short citation contexts, "
        f"{counts['excessive_dashes']} contexts with excessive dashes, "
        f"{counts['multiple_years']} self-cited references with multiple years."
    )


if __name__ == "__main__":
    cli()
