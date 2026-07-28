"""Convert an annotated Excel workbook to OpenAI fine-tuning JSONL files."""

from __future__ import annotations

import argparse
import json
import random
import re
from pathlib import Path
from typing import Sequence

import openpyxl


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_LONG_PROMPT = SCRIPT_DIR / "prompts" / "promptsprompt_long.txt"
DEFAULT_SHORT_PROMPT = SCRIPT_DIR / "prompts" / "promptsprompt_short.txt"

FUNCTION_LABELS = {
    "Foundation",
    "Inspiration",
    "Extension",
    "Application",
    "Elaborated Citation",
    "Comparison",
    "Similarity",
    "Affirmation",
    "Related work",
    "Simple mention",
    "Comparison between Related Work",
    "Future work",
    "Further reading",
    "Historical background",
}
DEPTH_LABELS = {"Deep citation", "Moderate citation", "Shallow citation"}

HEADER_ALIASES = {
    "article_index": (
        "Self-citing Article Index",
        "Self-cited Article Index",
    ),
    "article_title": ("Self-cited Article Title",),
    "citation_content": ("Citation Content",),
    "citation_location": ("citation location", "Citation Location"),
    "citation_function": ("citation function", "Citation Function"),
    "citation_depth": ("citation depth", "Citation Depth"),
}


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert annotated .xlsx rows to chat fine-tuning JSONL."
    )
    parser.add_argument("--input", required=True, type=Path, help="Annotated .xlsx.")
    parser.add_argument(
        "--train-output",
        required=True,
        type=Path,
        help="Destination for training JSONL.",
    )
    parser.add_argument(
        "--validation-output",
        type=Path,
        help="Destination for validation JSONL when --validation-ratio is positive.",
    )
    parser.add_argument(
        "--validation-ratio",
        type=float,
        default=0.0,
        help="Fraction assigned to validation, from 0 to less than 1 (default: 0).",
    )
    parser.add_argument("--sheet", help="Worksheet name; defaults to the active sheet.")
    parser.add_argument("--seed", type=int, default=42, help="Random seed (default: 42).")
    parser.add_argument(
        "--long-prompt-rate",
        type=float,
        default=0.5,
        help="Probability of using the long system prompt (default: 0.5).",
    )
    parser.add_argument(
        "--long-prompt",
        type=Path,
        default=DEFAULT_LONG_PROMPT,
        help="Long system-prompt text file.",
    )
    parser.add_argument(
        "--short-prompt",
        type=Path,
        default=DEFAULT_SHORT_PROMPT,
        help="Short system-prompt text file.",
    )
    parser.add_argument(
        "--no-shuffle",
        action="store_true",
        help="Preserve Excel row order instead of deterministic shuffling.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Replace output files if they already exist.",
    )
    args = parser.parse_args(argv)

    if not 0 <= args.validation_ratio < 1:
        parser.error("--validation-ratio must be at least 0 and less than 1")
    if not 0 <= args.long_prompt_rate <= 1:
        parser.error("--long-prompt-rate must be between 0 and 1")
    if args.validation_ratio == 0 and args.validation_output is not None:
        parser.error(
            "--validation-output requires a positive --validation-ratio"
        )
    return args


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value).strip()).casefold()


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def resolve_columns(header: tuple[object, ...]) -> dict[str, int]:
    normalized_to_indexes: dict[str, list[int]] = {}
    for index, value in enumerate(header):
        if value is None:
            continue
        normalized_to_indexes.setdefault(normalize_header(value), []).append(index)

    resolved: dict[str, int] = {}
    missing: list[str] = []
    for field, aliases in HEADER_ALIASES.items():
        matches = {
            index
            for alias in aliases
            for index in normalized_to_indexes.get(normalize_header(alias), [])
        }
        if not matches:
            missing.append(f"{field}: one of {list(aliases)}")
        elif len(matches) > 1:
            raise ValueError(
                f"Multiple columns match {field!r}: "
                f"{[header[index] for index in sorted(matches)]}"
            )
        else:
            resolved[field] = matches.pop()

    if missing:
        available = [text(value) for value in header if value is not None]
        raise ValueError(
            "Required columns are missing:\n- "
            + "\n- ".join(missing)
            + f"\nAvailable columns: {available}"
        )
    return resolved


def load_prompt(path: Path) -> str:
    resolved = path.expanduser().resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"Prompt file not found: {resolved}")
    prompt = resolved.read_text(encoding="utf-8").strip()
    if not prompt:
        raise ValueError(f"Prompt file is empty: {resolved}")
    return prompt


def build_record(
    row: tuple[object, ...],
    columns: dict[str, int],
    system_prompt: str,
    row_number: int,
) -> dict[str, object]:
    article_index = text(row[columns["article_index"]])
    article_title = text(row[columns["article_title"]])
    citation_content = text(row[columns["citation_content"]])
    citation_location = text(row[columns["citation_location"]])
    citation_function = text(row[columns["citation_function"]])
    citation_depth = text(row[columns["citation_depth"]])

    required_text = {
        "Self-cited Article Title": article_title,
        "Citation Content": citation_content,
        "Citation Location": citation_location,
    }
    blank = [name for name, value in required_text.items() if not value]
    if blank:
        raise ValueError(f"Excel row {row_number}: blank required fields: {blank}")
    if citation_function not in FUNCTION_LABELS:
        raise ValueError(
            f"Excel row {row_number}: invalid Citation Function "
            f"{citation_function!r}"
        )
    if citation_depth not in DEPTH_LABELS:
        raise ValueError(
            f"Excel row {row_number}: invalid Citation Depth {citation_depth!r}"
        )

    user_prompt = (
        f"Self-cited Article Index: {article_index}\n"
        f"Self-cited Article Title: {article_title}\n"
        f"Citation Location: {citation_location}\n"
        f"Citation Content: {citation_content}"
    )
    assistant_answer = json.dumps(
        {
            "Final Citation Function": citation_function,
            "Final Citation Depth": citation_depth,
        },
        ensure_ascii=False,
    )
    return {
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
            {"role": "assistant", "content": assistant_answer},
        ]
    }


def choose_validation_output(args: argparse.Namespace) -> Path | None:
    if args.validation_ratio == 0:
        return None
    if args.validation_output is not None:
        return args.validation_output.expanduser().resolve()
    train_output = args.train_output.expanduser().resolve()
    return train_output.with_name(
        f"{train_output.stem}_validation{train_output.suffix}"
    )


def check_output(path: Path, force: bool) -> None:
    if path.suffix.casefold() != ".jsonl":
        raise ValueError(f"Output must use the .jsonl extension: {path}")
    if path.exists() and not force:
        raise FileExistsError(
            f"Output already exists: {path}. Pass --force to replace it."
        )


def write_jsonl(path: Path, records: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8", newline="\n") as stream:
        for record in records:
            stream.write(json.dumps(record, ensure_ascii=False) + "\n")
    temporary.replace(path)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    input_path = args.input.expanduser().resolve()
    if not input_path.is_file():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")
    if input_path.suffix.casefold() != ".xlsx":
        raise ValueError(f"Input must use the .xlsx extension: {input_path}")

    train_output = args.train_output.expanduser().resolve()
    validation_output = choose_validation_output(args)
    check_output(train_output, args.force)
    if validation_output is not None:
        check_output(validation_output, args.force)
        if validation_output == train_output:
            raise ValueError("Training and validation outputs must be different.")

    long_prompt = load_prompt(args.long_prompt)
    short_prompt = load_prompt(args.short_prompt)
    rng = random.Random(args.seed)

    workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    if args.sheet is None:
        worksheet = workbook.active
    elif args.sheet in workbook.sheetnames:
        worksheet = workbook[args.sheet]
    else:
        raise ValueError(
            f"Worksheet {args.sheet!r} not found; available: {workbook.sheetnames}"
        )

    rows = worksheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration as exc:
        raise ValueError("Input worksheet is empty.") from exc
    columns = resolve_columns(header)

    records: list[dict[str, object]] = []
    for row_number, row in enumerate(rows, start=2):
        if not any(value is not None and text(value) for value in row):
            continue
        system_prompt = (
            long_prompt if rng.random() < args.long_prompt_rate else short_prompt
        )
        records.append(build_record(row, columns, system_prompt, row_number))
    workbook.close()

    if not records:
        raise ValueError("No valid data rows were found.")
    if not args.no_shuffle:
        rng.shuffle(records)

    validation_count = 0
    if args.validation_ratio > 0:
        validation_count = max(1, round(len(records) * args.validation_ratio))
        if validation_count >= len(records):
            raise ValueError("Validation split would leave no training records.")

    validation_records = records[:validation_count]
    training_records = records[validation_count:]
    write_jsonl(train_output, training_records)
    if validation_output is not None:
        write_jsonl(validation_output, validation_records)

    print(f"Input: {input_path}")
    print(f"Worksheet: {worksheet.title}")
    print(f"Seed: {args.seed}")
    print(f"Training records: {len(training_records)} -> {train_output}")
    if validation_output is not None:
        print(
            f"Validation records: {len(validation_records)} -> "
            f"{validation_output}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
