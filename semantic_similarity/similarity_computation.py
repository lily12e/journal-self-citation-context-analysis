"""Compute BERT-whitening semantic similarity for rows in an Excel workbook.

The default ``reported`` mode reproduces the semantic-similarity values
reported in the study. An optional ``corrected`` mode is retained for
methodological comparison.
"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

import numpy as np
import openpyxl
import torch
from openpyxl.utils import column_index_from_string, get_column_letter

from all_utils import load_whiten, transform_and_normalize


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_WHITENING = (
    SCRIPT_DIR
    / "resources"
    / "bert-base-uncased-first_last_avg-whiten(NLI).pkl"
)
EXCEL_COLUMN_RE = re.compile(r"^[A-Za-z]{1,3}$")


@dataclass(frozen=True)
class ColumnSelection:
    abstract: int
    citation: int
    output: int


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compute semantic similarity with BERT whitening."
    )
    parser.add_argument("--input", required=True, type=Path, help="Input .xlsx file.")
    parser.add_argument(
        "--output",
        type=Path,
        help="Output .xlsx file. Defaults to INPUT_with_similarity.xlsx.",
    )
    parser.add_argument(
        "--in-place",
        action="store_true",
        help="Overwrite the input workbook. Cannot be combined with --output.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Allow replacement of an existing output file.",
    )
    parser.add_argument(
        "--sheet",
        help="Worksheet name. Defaults to the active worksheet.",
    )
    parser.add_argument(
        "--abstract-column",
        default="Self-cited article abstract",
        help="Header text or Excel column letter for the cited-article abstract.",
    )
    parser.add_argument(
        "--citation-column",
        default="Citation content",
        help="Header text or Excel column letter for citation context.",
    )
    parser.add_argument(
        "--output-column",
        default="Semantic similarity",
        help="Existing header text or Excel column letter for output.",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=1,
        help="One-based header row number (default: 1).",
    )
    parser.add_argument(
        "--model",
        default="bert-base-uncased",
        help="Hugging Face model ID or a local model directory.",
    )
    parser.add_argument(
        "--model-revision",
        help="Optional Hugging Face model revision/commit.",
    )
    parser.add_argument(
        "--local-files-only",
        action="store_true",
        help="Do not download model files.",
    )
    parser.add_argument(
        "--whitening",
        type=Path,
        default=DEFAULT_WHITENING,
        help="Whitening parameter file.",
    )
    parser.add_argument(
        "--mode",
        choices=("reported", "corrected"),
        default="reported",
        help="Reproduce reported values (default) or use corrected vectors.",
    )
    parser.add_argument(
        "--pooling",
        choices=("auto", "mean", "cls"),
        default="auto",
        help="'auto' selects CLS for reported mode and mean for corrected mode.",
    )
    parser.add_argument(
        "--device",
        choices=("auto", "cpu", "cuda"),
        default="auto",
    )
    parser.add_argument(
        "--chunk-tokens",
        type=int,
        default=450,
        help="Approximate maximum tokens per text chunk (default: 450).",
    )
    parser.add_argument(
        "--limit",
        type=int,
        help="Process at most this many data rows (useful for a smoke test).",
    )
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Validate workbook, columns, and resources without loading BERT.",
    )
    args = parser.parse_args(argv)

    if args.in_place and args.output is not None:
        parser.error("--in-place cannot be combined with --output")
    if args.in_place and args.force:
        parser.error("--force is unnecessary with the explicit --in-place option")
    if args.header_row < 1:
        parser.error("--header-row must be at least 1")
    if args.chunk_tokens < 1 or args.chunk_tokens > 510:
        parser.error("--chunk-tokens must be between 1 and 510")
    if args.limit is not None and args.limit < 1:
        parser.error("--limit must be at least 1")
    return args


def choose_device(requested: str) -> torch.device:
    if requested == "cuda":
        if not torch.cuda.is_available():
            raise RuntimeError("--device cuda was requested, but CUDA is unavailable.")
        return torch.device("cuda")
    if requested == "cpu":
        return torch.device("cpu")
    return torch.device("cuda" if torch.cuda.is_available() else "cpu")


def normalized_header(value: object) -> str:
    return " ".join(str(value).strip().casefold().split())


def resolve_column(worksheet, selector: str, header_row: int) -> int:
    selector = selector.strip()
    if EXCEL_COLUMN_RE.fullmatch(selector):
        return column_index_from_string(selector.upper())

    wanted = normalized_header(selector)
    matches = [
        cell.column
        for cell in worksheet[header_row]
        if cell.value is not None and normalized_header(cell.value) == wanted
    ]
    if not matches:
        available = [
            str(cell.value).strip()
            for cell in worksheet[header_row]
            if cell.value is not None
        ]
        raise ValueError(
            f"Column header not found: {selector!r}. Available headers: {available}"
        )
    if len(matches) > 1:
        raise ValueError(f"Column header is duplicated: {selector!r}")
    return int(matches[0])


def select_worksheet(workbook, name: str | None):
    if name is None:
        return workbook.active
    if name not in workbook.sheetnames:
        raise ValueError(
            f"Worksheet {name!r} not found. Available sheets: {workbook.sheetnames}"
        )
    return workbook[name]


def resolve_output_path(args: argparse.Namespace) -> Path:
    input_path = args.input.expanduser().resolve()
    if args.in_place:
        return input_path
    if args.output is not None:
        return args.output.expanduser().resolve()
    return input_path.with_name(f"{input_path.stem}_with_similarity{input_path.suffix}")


class SimilarityCalculator:
    def __init__(
        self,
        *,
        model_name: str,
        model_revision: str | None,
        local_files_only: bool,
        whitening_path: Path,
        device: torch.device,
        mode: str,
        pooling: str,
        chunk_tokens: int,
    ) -> None:
        try:
            from transformers import AutoModel, AutoTokenizer
        except ImportError as exc:
            raise RuntimeError(
                "The 'transformers' package is required for similarity "
                "calculation. Install semantic_similarity/requirements.txt."
            ) from exc

        self.device = device
        self.mode = mode
        self.pooling = (
            ("mean" if mode == "corrected" else "cls")
            if pooling == "auto"
            else pooling
        )
        self.chunk_tokens = chunk_tokens

        model_kwargs = {
            "revision": model_revision,
            "local_files_only": local_files_only,
        }
        self.tokenizer = AutoTokenizer.from_pretrained(model_name, **model_kwargs)
        self.model = AutoModel.from_pretrained(model_name, **model_kwargs)
        self.model.to(self.device)
        self.model.eval()

        self.kernel, self.bias = load_whiten(whitening_path)
        hidden_size = int(self.model.config.hidden_size)
        if self.kernel.shape[0] != hidden_size:
            raise ValueError(
                "Model and whitening dimensions do not match: "
                f"model={hidden_size}, kernel={self.kernel.shape}"
            )

    def chunk_text(self, text: str) -> list[str]:
        sentences = text.split(". ")
        if len(sentences) == 1:
            units = text.split()
            separator = " "
            suffix = ""
        else:
            units = sentences
            separator = ". "
            suffix = "."

        chunks: list[str] = []
        current: list[str] = []
        current_length = 0
        for unit in units:
            unit_length = len(self.tokenizer.tokenize(unit))
            if current and current_length + unit_length > self.chunk_tokens:
                chunks.append(separator.join(current) + suffix)
                current = [unit]
                current_length = unit_length
            else:
                current.append(unit)
                current_length += unit_length
        if current:
            chunk = separator.join(current)
            if suffix and not chunk.endswith("."):
                chunk += suffix
            chunks.append(chunk)
        return chunks or [text[:1000]]

    def embed(self, text: str) -> torch.Tensor:
        chunk_embeddings: list[torch.Tensor] = []
        for chunk in self.chunk_text(text):
            if not chunk.strip():
                continue
            inputs = self.tokenizer(
                chunk,
                return_tensors="pt",
                truncation=True,
                max_length=512,
            )
            inputs = {key: value.to(self.device) for key, value in inputs.items()}
            with torch.inference_mode():
                outputs = self.model(**inputs, output_hidden_states=True)

            first = outputs.hidden_states[1]
            last = outputs.hidden_states[-1]
            if self.pooling == "cls":
                vector = (first[:, 0, :] + last[:, 0, :]) / 2
            else:
                vector = (first + last).mean(dim=1)
            chunk_embeddings.append(vector.squeeze(0))

        if not chunk_embeddings:
            raise ValueError("Text produced no usable chunks.")
        return torch.stack(chunk_embeddings).mean(dim=0)

    def similarity(self, text1: str, text2: str) -> float:
        vectors = torch.stack([self.embed(text1), self.embed(text2)])
        if self.mode == "reported":
            # Tensor layout used to generate the values reported in the study.
            vectors = vectors[:, None, None, :]

        transformed = transform_and_normalize(
            vectors,
            kernel=self.kernel,
            bias=self.bias,
            normalization_dim=1,
        )
        first = transformed[0].flatten()
        second = transformed[1].flatten()
        return float(
            torch.nn.functional.cosine_similarity(first, second, dim=0)
            .detach()
            .cpu()
        )


def validate_inputs(
    args: argparse.Namespace,
) -> tuple[Path, Path, object, object, ColumnSelection]:
    input_path = args.input.expanduser().resolve()
    if not input_path.is_file():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")
    if input_path.suffix.casefold() != ".xlsx":
        raise ValueError("Input must be an .xlsx workbook.")

    whitening_path = args.whitening.expanduser().resolve()
    load_whiten(whitening_path)

    workbook = openpyxl.load_workbook(input_path, read_only=args.validate_only)
    worksheet = select_worksheet(workbook, args.sheet)
    columns = ColumnSelection(
        abstract=resolve_column(worksheet, args.abstract_column, args.header_row),
        citation=resolve_column(worksheet, args.citation_column, args.header_row),
        output=resolve_column(worksheet, args.output_column, args.header_row),
    )
    output_path = resolve_output_path(args)
    return input_path, output_path, workbook, worksheet, columns


def process_workbook(args: argparse.Namespace) -> Path | None:
    input_path, output_path, workbook, worksheet, columns = validate_inputs(args)
    print(f"Input: {input_path}")
    print(f"Worksheet: {worksheet.title}")
    print(
        "Columns: "
        f"abstract={get_column_letter(columns.abstract)}, "
        f"citation={get_column_letter(columns.citation)}, "
        f"output={get_column_letter(columns.output)}"
    )
    print(f"Whitening: {args.whitening.expanduser().resolve()}")

    if args.validate_only:
        workbook.close()
        print("Validation passed. The model was not loaded and no file was written.")
        return None

    if output_path != input_path:
        if output_path.exists() and not args.force:
            raise FileExistsError(
                f"Output already exists: {output_path}. Pass --force to replace it."
            )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.close()
        shutil.copy2(input_path, output_path)
        workbook = openpyxl.load_workbook(output_path)
        worksheet = select_worksheet(workbook, args.sheet)

    calculator = SimilarityCalculator(
        model_name=args.model,
        model_revision=args.model_revision,
        local_files_only=args.local_files_only,
        whitening_path=args.whitening.expanduser().resolve(),
        device=choose_device(args.device),
        mode=args.mode,
        pooling=args.pooling,
        chunk_tokens=args.chunk_tokens,
    )
    print(
        f"Mode: {args.mode}; pooling: {calculator.pooling}; "
        f"device: {calculator.device}"
    )

    start_row = args.header_row + 1
    stop_row = worksheet.max_row
    if args.limit is not None:
        stop_row = min(stop_row, start_row + args.limit - 1)

    processed = skipped = failed = 0
    started = time.monotonic()
    for row_number in range(start_row, stop_row + 1):
        abstract = worksheet.cell(row_number, columns.abstract).value
        citation = worksheet.cell(row_number, columns.citation).value
        if abstract is None or citation is None:
            skipped += 1
            continue
        abstract_text = str(abstract).strip()
        citation_text = str(citation).strip()
        if not abstract_text or not citation_text:
            skipped += 1
            continue

        try:
            score = calculator.similarity(abstract_text, citation_text)
        except Exception as exc:
            failed += 1
            print(f"Row {row_number} failed: {exc}", file=sys.stderr)
            continue
        worksheet.cell(row_number, columns.output).value = score
        processed += 1
        if processed % 100 == 0:
            elapsed = time.monotonic() - started
            print(f"Processed {processed} rows in {elapsed:.1f} seconds.")

    workbook.save(output_path)
    workbook.close()
    print(
        f"Done: processed={processed}, skipped={skipped}, failed={failed}; "
        f"output={output_path}"
    )
    if failed:
        raise RuntimeError(f"{failed} row(s) failed; inspect the messages above.")
    return output_path


def main(argv: Sequence[str] | None = None) -> int:
    try:
        args = parse_args(argv)
        process_workbook(args)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
